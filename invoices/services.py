import openpyxl
from decimal import Decimal, InvalidOperation
from django.db import transaction
from .models import ProductInventory, ServiceType

def ingest_inventory_excel(file_stream):
    try:
        workbook = openpyxl.load_workbook(file_stream, data_only=True)
        sheet = workbook.active
        
        header = [str(cell.value).strip().lower() for cell in sheet[1]]
        required_cols = ['category', 'model number', 'part name', 'part code', 'price']
        
        for col in required_cols:
            if col not in header:
                return {'success': False, 'error': f"Missing column: {col}"}

        indices = {col: header.index(col) for col in required_cols}
        processed = updated = created = 0

        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                category = row[indices['category']]
                model_number = row[indices['model number']]
                part_name = row[indices['part name']]
                part_code = row[indices['part code']]
                price_raw = row[indices['price']]

                if not all([category, model_number, part_name, part_code]):
                    continue

                try:
                    price_pkr = Decimal(str(price_raw))
                except (InvalidOperation, TypeError, ValueError):
                    continue

                obj, is_new = ProductInventory.objects.update_or_create(
                    part_code=str(part_code).strip(),
                    defaults={
                        'category': str(category).strip(),
                        'model_number': str(model_number).strip(),
                        'part_name': str(part_name).strip(),
                        'price_pkr': price_pkr,
                    }
                )
                
                processed += 1
                if is_new:
                    created += 1
                else:
                    updated += 1
                    
        return {'success': True, 'processed': processed, 'created': created, 'updated': updated}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}

def ingest_services_excel(file_stream):
    try:
        workbook = openpyxl.load_workbook(file_stream, data_only=True)
        sheet = workbook.active
        
        header = [str(cell.value).strip().lower() for cell in sheet[1]]
        required_cols = ['service name', 'description', 'worker fee']
        
        for col in required_cols:
            if col not in header:
                return {'success': False, 'error': f"Missing column: {col}"}

        indices = {col: header.index(col) for col in required_cols}
        processed = updated = created = 0

        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                name = row[indices['service name']]
                desc = row[indices['description']] or ""
                fee_raw = row[indices['worker fee']]

                if not name or fee_raw is None:
                    continue

                try:
                    fee_pkr = Decimal(str(fee_raw))
                except (InvalidOperation, TypeError, ValueError):
                    continue

                obj, is_new = ServiceType.objects.update_or_create(
                    name=str(name).strip(),
                    defaults={
                        'description': str(desc).strip(),
                        'worker_fee_pkr': fee_pkr,
                    }
                )
                
                processed += 1
                if is_new:
                    created += 1
                else:
                    updated += 1
                    
        return {'success': True, 'processed': processed, 'created': created, 'updated': updated}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}
